from __future__ import unicode_literals
from six import StringIO, text_type
from uuid import UUID

from ordereddict import OrderedDict
from openpyxl import Workbook

from rest_framework.renderers import BaseRenderer

# six versions 1.3.0 and previous don't have PY2
try:
    from six import PY2
except ImportError:
    import sys
    PY2 = sys.version_info[0] == 2


class ExcelRenderer(BaseRenderer):

    """
    Renderer which serializes to Excel
    """

    media_type = 'application/ms-excel'
    format = 'xlsx'
    level_sep = '.'

    def render(self, data, media_type=None, renderer_context=None):
        """
        Renders serialized *data* into Excel. For a dictionary:
        """
        if data is None:
            return ''

        stream = StringIO()

        wb = Workbook()
        ws = wb.active

        table = self.tablize(data)

        for row_idx, row in enumerate(table, 1):
            for col_idx, elem in enumerate(row, 1):
                if isinstance(elem, text_type) and PY2:
                    ws.cell(column=col_idx, row=row_idx, value=elem.encode('utf-8'))
                elif isinstance(elem, UUID):
                    ws.cell(column=col_idx, row=row_idx, value=unicode(elem))
                else:
                    ws.cell(column=col_idx, row=row_idx, value=elem)

        wb.save(stream)

        return stream.getvalue()

    def tablize(self, data):
        """
        Convert a list of data into a table.
        """
        if data:

            # First, flatten the data (i.e., convert it to a list of
            # dictionaries that are each exactly one level deep).  The key for
            # each item designates the name of the column that the item will
            # fall into.
            data = self.flatten_data(data)

            # Get the set of all unique headers, and sort them.
            headers = []
            if data[0]:
                headers = data[0].keys()

            # Create a row for each dictionary, filling in columns for which the
            # item has no data with None values.
            rows = []
            for item in data:
                row = []
                for key in headers:
                    row.append(item.get(key, None))
                rows.append(row)

            # Return your "table", with the headers as the first row.
            return [headers] + rows

        else:

            return []

    def flatten_data(self, data):
        """
        Convert the given data collection to a list of dictionaries that are
        each exactly one level deep. The key for each value in the dictionaries
        designates the name of the column that the value will fall into.
        """
        flat_data = []
        for item in data:
            flat_item = self.flatten_item(item)
            flat_data.append(flat_item)

        return flat_data

    def flatten_item(self, item):
        if isinstance(item, list):
            flat_item = self.flatten_list(item)
        elif isinstance(item, dict):
            flat_item = self.flatten_dict(item)
        else:
            flat_item = {'': item}

        return flat_item

    def nest_flat_item(self, flat_item, prefix):
        """
        Given a "flat item" (a dictionary exactly one level deep), nest all of
        the column headers in a namespace designated by prefix.  For example:
         header... | with prefix... | becomes...
        -----------|----------------|----------------
         'lat'     | 'location'     | 'location.lat'
         ''        | '0'            | '0'
         'votes.1' | 'user'         | 'user.votes.1'
        """
        nested_item = {}
        for header, val in flat_item.items():
            nested_header = self.level_sep.join([prefix, header]) if header else prefix
            nested_item[nested_header] = val
        return nested_item

    def flatten_list(self, l):
        flat_list = {}
        for index, item in enumerate(l):
            index = text_type(index)
            flat_item = self.flatten_item(item)
            nested_item = self.nest_flat_item(flat_item, index)
            flat_list.update(nested_item)
        return flat_list

    def flatten_dict(self, d):
        flat_dict = OrderedDict()
        for key, item in d.iteritems():
            key = str(key)
            flat_item = self.flatten_item(item)
            nested_item = self.nest_flat_item(flat_item, key)
            flat_dict.update(nested_item)
        return flat_dict
